import customtkinter as ctk
from tkinter import messagebox, filedialog
from src.controllers.auth_controller import AuthController
from src.controllers.business_controllers import ProductController
from src.models import Product, UserRole


class ProductsView(ctk.CTkFrame):
    def __init__(self, parent, controller: ProductController, auth: AuthController):
        super().__init__(parent, fg_color="transparent")
        self.controller = controller
        self.auth = auth
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.selected_product = None
        self.setup_ui()
        self.refresh_list()
        
    def setup_ui(self):
        toolbar = ctk.CTkFrame(self, height=60, corner_radius=10)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        toolbar.grid_columnconfigure(1, weight=1)
        
        ctk.CTkButton(toolbar, text="➕ Novo Produto", height=40,
                     font=ctk.CTkFont(size=13, weight="bold"),
                     command=self.open_form).grid(row=0, column=0, padx=15, pady=10)
        
        self.search_entry = ctk.CTkEntry(toolbar, placeholder_text="🔍 Buscar por nome, SKU, modelo...", height=40,
                                        font=ctk.CTkFont(size=13), width=300)
        self.search_entry.grid(row=0, column=1, padx=10, pady=10, sticky="e")
        self.search_entry.bind("<KeyRelease>", lambda e: self.refresh_list())
        
        ctk.CTkButton(toolbar, text="📤 Exportar", height=40, width=100,
                     command=self.export_excel).grid(row=0, column=2, padx=15, pady=10)
        
        list_frame = ctk.CTkFrame(self, corner_radius=10)
        list_frame.grid(row=1, column=0, sticky="nsew")
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(0, weight=1)
        
        columns = ("ID", "SKU", "Produto", "Marca", "Categoria", "Estoque", "Mín", "Preço Venda", "Status")
        self.tree = ctk.CTkScrollableFrame(list_frame)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.tree.grid_columnconfigure(0, weight=1)
        
        self.header_frame = ctk.CTkFrame(self.tree, fg_color=("gray85", "gray20"))
        self.header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        col_weights = [0, 1, 3, 1, 1, 1, 1, 1, 1]
        for i, (col, weight) in enumerate(zip(columns, col_weights)):
            self.header_frame.grid_columnconfigure(i, weight=weight)
            ctk.CTkLabel(self.header_frame, text=col, font=ctk.CTkFont(weight="bold", size=12)).grid(
                row=0, column=i, padx=10, pady=8, sticky="w")
        
        self.rows_frame = ctk.CTkFrame(self.tree, fg_color="transparent")
        self.rows_frame.grid(row=1, column=0, sticky="nsew")
        self.rows_frame.grid_columnconfigure(0, weight=1)
        
        self.context_menu = ctk.CTkFrame(self, fg_color=("gray90", "gray15"), corner_radius=8, border_width=1)
        self.context_menu.grid_remove()
        ctk.CTkButton(self.context_menu, text="✏️ Editar", anchor="w", command=self.edit_selected).pack(fill="x", padx=5, pady=2)
        ctk.CTkButton(self.context_menu, text="📦 Ajustar Estoque", anchor="w", command=self.adjust_stock).pack(fill="x", padx=5, pady=2)
        ctk.CTkButton(self.context_menu, text="🗑️ Desativar", anchor="w", fg_color="#e74c3c", hover_color="#c0392b", command=self.delete_selected).pack(fill="x", padx=5, pady=2)
        
        self.bind("<Button-1>", lambda e: self.context_menu.grid_remove())
        
    def refresh_list(self):
        for widget in self.rows_frame.winfo_children():
            widget.destroy()
            
        term = self.search_entry.get().strip()
        products = self.controller.search_products(term) if term else self.controller.list_products()
        
        for idx, prod in enumerate(products):
            row = ctk.CTkFrame(self.rows_frame, fg_color=("gray90", "gray15") if idx % 2 == 0 else "transparent")
            row.grid(row=idx, column=0, sticky="ew", pady=1)
            row.grid_columnconfigure(0, weight=1)
            
            values = [
                str(prod.id), prod.sku, prod.name, prod.brand_name,
                prod.category_name, str(prod.current_stock), str(prod.min_stock),
                f"R$ {prod.sale_price:.2f}", "Ativo" if prod.is_active else "Inativo"
            ]
            
            for i, val in enumerate(values):
                lbl = ctk.CTkLabel(row, text=val, font=ctk.CTkFont(size=12), anchor="w")
                lbl.grid(row=0, column=i, padx=10, pady=8, sticky="w")
                self.header_frame.grid_columnconfigure(i, weight=col_weights[i])
                
            if prod.current_stock <= prod.min_stock and prod.is_active:
                for widget in row.winfo_children():
                    widget.configure(text_color="#e74c3c")
                    
            row.bind("<Button-3>", lambda e, p=prod: self.show_context_menu(e, p))
            row.bind("<Double-Button-1>", lambda e, p=prod: self.edit_product(p))
            for widget in row.winfo_children():
                widget.bind("<Button-3>", lambda e, p=prod: self.show_context_menu(e, p))
                widget.bind("<Double-Button-1>", lambda e, p=prod: self.edit_product(p))
                
    def show_context_menu(self, event, product):
        self.selected_product = product
        self.context_menu.grid(row=0, column=0)
        self.context_menu.place(x=event.x_root - self.winfo_rootx(), y=event.y_root - self.winfo_rooty())
        
    def open_form(self, product: Product = None):
        ProductForm(self, self.controller, product, callback=self.refresh_list)
        
    def edit_selected(self):
        if self.selected_product:
            self.open_form(self.selected_product)
        self.context_menu.grid_remove()
        
    def edit_product(self, product: Product):
        self.open_form(product)
        
    def adjust_stock(self):
        if not self.selected_product:
            return
        AdjustStockDialog(self, self.controller, self.selected_product, self.auth.current_user.id, self.refresh_list)
        self.context_menu.grid_remove()
        
    def delete_selected(self):
        if not self.selected_product:
            return
        if messagebox.askyesno("Confirmar", f"Desativar produto {self.selected_product.name}?"):
            success, msg = self.controller.delete_product(self.selected_product.id)
            messagebox.showinfo("Resultado", msg)
            self.refresh_list()
        self.context_menu.grid_remove()
        
    def export_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        products = self.controller.list_products(active_only=False)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos"
        
        headers = ["ID", "SKU", "Nome", "Marca", "Categoria", "Modelo", "Cor", "Armazenamento", "RAM",
                   "Tela", "Bateria", "Custo", "Venda", "Estoque", "Mínimo", "Ativo"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            
        for row, prod in enumerate(products, 2):
            ws.cell(row=row, column=1, value=prod.id)
            ws.cell(row=row, column=2, value=prod.sku)
            ws.cell(row=row, column=3, value=prod.name)
            ws.cell(row=row, column=4, value=prod.brand_name)
            ws.cell(row=row, column=5, value=prod.category_name)
            ws.cell(row=row, column=6, value=prod.model or "")
            ws.cell(row=row, column=7, value=prod.color or "")
            ws.cell(row=row, column=8, value=prod.storage_gb or "")
            ws.cell(row=row, column=9, value=prod.ram_gb or "")
            ws.cell(row=row, column=10, value=prod.screen_inches or "")
            ws.cell(row=row, column=11, value=prod.battery_mah or "")
            ws.cell(row=row, column=12, value=prod.cost_price)
            ws.cell(row=row, column=13, value=prod.sale_price)
            ws.cell(row=row, column=14, value=prod.current_stock)
            ws.cell(row=row, column=15, value=prod.min_stock)
            ws.cell(row=row, column=16, value="Sim" if prod.is_active else "Não")
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialname=f"produtos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Exportado para {file_path}")


class ProductForm(ctk.CTkToplevel):
    def __init__(self, parent, controller: ProductController, product: Product = None, callback=None):
        super().__init__(parent)
        self.controller = controller
        self.product = product
        self.callback = callback
        self.is_edit = product is not None
        
        self.title("Editar Produto" if self.is_edit else "Novo Produto")
        self.geometry("600x750")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.center_window()
        
        self.brands = controller.list_brands()
        self.categories = controller.list_categories()
        
        self.setup_ui()
        if self.is_edit:
            self.load_data()
            
    def center_window(self):
        self.update_idletasks()
        x = self.master.winfo_rootx() + (self.master.winfo_width() // 2) - 300
        y = self.master.winfo_rooty() + (self.master.winfo_height() // 2) - 375
        self.geometry(f"+{x}+{y}")
        
    def setup_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        scroll = ctk.CTkScrollableFrame(self)
        scroll.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        scroll.grid_columnconfigure(1, weight=1)
        
        fields = [
            ("sku", "SKU *", "entry", {}),
            ("name", "Nome *", "entry", {}),
            ("brand_id", "Marca *", "combo", {"values": [b.name for b in self.brands]}),
            ("category_id", "Categoria *", "combo", {"values": [c.name for c in self.categories]}),
            ("model", "Modelo", "entry", {}),
            ("color", "Cor", "entry", {}),
            ("storage_gb", "Armazenamento (GB)", "number", {}),
            ("ram_gb", "RAM (GB)", "number", {}),
            ("screen_inches", "Tela (pol)", "number", {"step": 0.1}),
            ("battery_mah", "Bateria (mAh)", "number", {}),
            ("cost_price", "Preço de Custo *", "money", {}),
            ("sale_price", "Preço de Venda *", "money", {}),
            ("min_stock", "Estoque Mínimo", "number", {}),
            ("current_stock", "Estoque Atual", "number", {}),
            ("description", "Descrição", "text", {}),
        ]
        
        self.widgets = {}
        for row, (field, label, ftype, opts) in enumerate(fields):
            ctk.CTkLabel(scroll, text=label, font=ctk.CTkFont(size=13)).grid(row=row, column=0, padx=10, pady=8, sticky="w")
            
            if ftype == "entry":
                w = ctk.CTkEntry(scroll, font=ctk.CTkFont(size=13), height=35)
            elif ftype == "combo":
                w = ctk.CTkComboBox(scroll, values=opts["values"], font=ctk.CTkFont(size=13), height=35)
            elif ftype == "number":
                w = ctk.CTkEntry(scroll, font=ctk.CTkFont(size=13), height=35, placeholder_text="0")
            elif ftype == "money":
                w = ctk.CTkEntry(scroll, font=ctk.CTkFont(size=13), height=35, placeholder_text="0.00")
            elif ftype == "text":
                w = ctk.CTkTextbox(scroll, font=ctk.CTkFont(size=13), height=80)
                
            w.grid(row=row, column=1, padx=10, pady=8, sticky="ew")
            self.widgets[field] = w
            
        self.active_switch = ctk.CTkSwitch(scroll, text="Produto Ativo", font=ctk.CTkFont(size=13))
        self.active_switch.grid(row=len(fields), column=0, columnspan=2, padx=10, pady=15, sticky="w")
        self.active_switch.select()
        
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.grid(row=len(fields)+1, column=0, columnspan=2, pady=20)
        btn_frame.grid_columnconfigure((0,1), weight=1)
        
        ctk.CTkButton(btn_frame, text="Cancelar", height=40, fg_color="gray", hover_color="gray30",
                     command=self.destroy).grid(row=0, column=0, padx=10, sticky="ew")
        ctk.CTkButton(btn_frame, text="Salvar", height=40,
                     command=self.save).grid(row=0, column=1, padx=10, sticky="ew")
        
    def load_data(self):
        p = self.product
        self.widgets["sku"].insert(0, p.sku)
        self.widgets["sku"].configure(state="disabled")
        self.widgets["name"].insert(0, p.name)
        self.widgets["brand_id"].set(p.brand_name)
        self.widgets["category_id"].set(p.category_name)
        if p.model: self.widgets["model"].insert(0, p.model)
        if p.color: self.widgets["color"].insert(0, p.color)
        if p.storage_gb: self.widgets["storage_gb"].insert(0, str(p.storage_gb))
        if p.ram_gb: self.widgets["ram_gb"].insert(0, str(p.ram_gb))
        if p.screen_inches: self.widgets["screen_inches"].insert(0, str(p.screen_inches))
        if p.battery_mah: self.widgets["battery_mah"].insert(0, str(p.battery_mah))
        self.widgets["cost_price"].insert(0, f"{p.cost_price:.2f}")
        self.widgets["sale_price"].insert(0, f"{p.sale_price:.2f}")
        self.widgets["min_stock"].insert(0, str(p.min_stock))
        self.widgets["current_stock"].insert(0, str(p.current_stock))
        if p.description: self.widgets["description"].insert("1.0", p.description)
        if p.is_active: self.active_switch.select()
        else: self.active_switch.deselect()
        
    def save(self):
        try:
            brand_name = self.widgets["brand_id"].get()
            cat_name = self.widgets["category_id"].get()
            brand = next((b for b in self.brands if b.name == brand_name), None)
            cat = next((c for c in self.categories if c.name == cat_name), None)
            
            if not brand or not cat:
                messagebox.showerror("Erro", "Selecione marca e categoria válidas")
                return
                
            product = Product(
                id=self.product.id if self.is_edit else None,
                sku=self.widgets["sku"].get().strip(),
                name=self.widgets["name"].get().strip(),
                brand_id=brand.id,
                category_id=cat.id,
                model=self.widgets["model"].get().strip() or None,
                color=self.widgets["color"].get().strip() or None,
                storage_gb=int(self.widgets["storage_gb"].get()) if self.widgets["storage_gb"].get() else None,
                ram_gb=int(self.widgets["ram_gb"].get()) if self.widgets["ram_gb"].get() else None,
                screen_inches=float(self.widgets["screen_inches"].get()) if self.widgets["screen_inches"].get() else None,
                battery_mah=int(self.widgets["battery_mah"].get()) if self.widgets["battery_mah"].get() else None,
                description=self.widgets["description"].get("1.0", "end-1c").strip() or None,
                cost_price=float(self.widgets["cost_price"].get() or 0),
                sale_price=float(self.widgets["sale_price"].get() or 0),
                min_stock=int(self.widgets["min_stock"].get() or 5),
                current_stock=int(self.widgets["current_stock"].get() or 0),
                is_active=self.active_switch.get() == 1
            )
            
            if self.is_edit:
                success, msg = self.controller.update_product(product)
            else:
                success, msg, _ = self.controller.create_product(product)
                
            if success:
                messagebox.showinfo("Sucesso", msg)
                if self.callback:
                    self.callback()
                self.destroy()
            else:
                messagebox.showerror("Erro", msg)
        except ValueError as e:
            messagebox.showerror("Erro", "Valores numéricos inválidos")
        except Exception as e:
            messagebox.showerror("Erro", str(e))


class AdjustStockDialog(ctk.CTkToplevel):
    def __init__(self, parent, controller: ProductController, product: Product, user_id: int, callback):
        super().__init__(parent)
        self.controller = controller
        self.product = product
        self.user_id = user_id
        self.callback = callback
        
        self.title(f"Ajustar Estoque: {product.name}")
        self.geometry("400x300")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.center_window()
        
        self.setup_ui()
        
    def center_window(self):
        self.update_idletasks()
        x = self.master.winfo_rootx() + (self.master.winfo_width() // 2) - 200
        y = self.master.winfo_rooty() + (self.master.winfo_height() // 2) - 150
        self.geometry(f"+{x}+{y}")
        
    def setup_ui(self):
        self.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(self, text=f"Produto: {self.product.name}", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=20)
        ctk.CTkLabel(self, text=f"Estoque Atual: {self.product.current_stock}", font=ctk.CTkFont(size=13)).pack(pady=5)
        ctk.CTkLabel(self, text=f"Estoque Mínimo: {self.product.min_stock}", font=ctk.CTkFont(size=13)).pack(pady=5)
        
        ctk.CTkLabel(self, text="Nova Quantidade:", font=ctk.CTkFont(size=13)).pack(pady=(20, 5))
        self.qty_entry = ctk.CTkEntry(self, font=ctk.CTkFont(size=16), height=45, justify="center")
        self.qty_entry.pack(padx=40, fill="x")
        self.qty_entry.insert(0, str(self.product.current_stock))
        self.qty_entry.focus()
        
        ctk.CTkLabel(self, text="Motivo:", font=ctk.CTkFont(size=13)).pack(pady=(15, 5))
        self.notes_entry = ctk.CTkEntry(self, font=ctk.CTkFont(size=13), height=35, placeholder_text="Ex: Ajuste inventário")
        self.notes_entry.pack(padx=40, fill="x")
        
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=20, fill="x", padx=40)
        btn_frame.grid_columnconfigure((0,1), weight=1)
        
        ctk.CTkButton(btn_frame, text="Cancelar", height=40, fg_color="gray", command=self.destroy).grid(row=0, column=0, padx=5, sticky="ew")
        ctk.CTkButton(btn_frame, text="Confirmar", height=40, command=self.confirm).grid(row=0, column=1, padx=5, sticky="ew")
        
    def confirm(self):
        try:
            new_qty = int(self.qty_entry.get())
            notes = self.notes_entry.get().strip()
            success, msg = self.controller.adjust_stock(self.product.id, new_qty, self.user_id, notes)
            if success:
                messagebox.showinfo("Sucesso", msg)
                if self.callback:
                    self.callback()
                self.destroy()
            else:
                messagebox.showerror("Erro", msg)
        except ValueError:
            messagebox.showerror("Erro", "Quantidade inválida")